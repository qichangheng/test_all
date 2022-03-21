#!/usr/bin/env python3
import openpyxl
from openpyxl import load_workbook

workbook = load_workbook('cases.xlsx')

from openpyxl.worksheet.worksheet import Worksheet

# 第一步：打开文件


workbook = openpyxl.load_workbook('cases.xlsx')
print(workbook)

# workbook.worksheets

# 第二步：获取表单
sheet: Worksheet = workbook['Sheet1']
print(sheet)

# 获取某一行，某一列的单元格
# cell() 方法得到的是一个单元格对象，不是 case_id 值
cell = sheet.cell(row=2, column=3)
print(cell)
# 获取单元格的数据
print(cell.value)


# 写入操作， 直接对 cell 当中value 属性赋值。
print(cell.value)
print(sheet.cell(row=2, column=3).value)

cell.value = 'login failed'
print(cell.value)

# 保存修改， 工作博.save("文件路径")
workbook.save('cases.xlsx')
# 关闭文件
workbook.close()

